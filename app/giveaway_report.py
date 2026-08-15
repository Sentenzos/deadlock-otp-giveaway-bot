from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
import math
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


MOSCOW_TIMEZONE = timezone(timedelta(hours=3), name="МСК")
RESULT_WINNER = "Победитель"
RESULT_NOT_SELECTED = "Не выбран"
RESULT_NOT_ELIGIBLE = "Не допущен"
ALLOWED_RESULTS = frozenset(
    {RESULT_WINNER, RESULT_NOT_SELECTED, RESULT_NOT_ELIGIBLE}
)

_MISSING = object()
_ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_DECIMAL_IDENTIFIER = re.compile(r"[0-9]+")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_EXACT_NUMERIC_SCORE_MAX = 999_999_999_999

_NAVY = "1F4E78"
_BLUE = "2F75B5"
_PALE_BLUE = "D9EAF7"
_LIGHT_BLUE = "EAF3F8"
_WHITE = "FFFFFF"
_TEXT = "1F2937"
_MUTED = "64748B"
_GREEN = "E2F0D9"
_GREEN_TEXT = "375623"
_RED = "FCE4D6"
_RED_TEXT = "9C0006"
_YELLOW = "FFF2CC"
_YELLOW_TEXT = "7F6000"
_BORDER = "B4C7D6"

_THIN_BORDER = Border(
    bottom=Side(style="thin", color=_BORDER),
)


@dataclass(frozen=True, slots=True)
class GiveawayReportMetadata:
    giveaway_id: int
    title: str
    prize: str
    min_minutes: float
    min_messages: int
    winner_count: int
    min_participants: int
    message_interval_seconds: int
    started_at: int | float | datetime | str | None = None
    end_at: int | float | datetime | str | None = None
    finished_at: int | float | datetime | str | None = None
    eligible_count: int | None = None


@dataclass(frozen=True, slots=True)
class GiveawayReportParticipant:
    twitch_login: str
    minutes: float
    messages: int
    admitted: bool
    random_number: str | None
    eligible_place: int | None
    result: str


@dataclass(frozen=True, slots=True)
class GiveawayReportRound:
    draw_round: int
    forced: bool
    participants: Sequence[GiveawayReportParticipant]


def build_giveaway_report(
    metadata: GiveawayReportMetadata | Mapping[str, Any] | object,
    rounds: Iterable[GiveawayReportRound | Mapping[str, Any] | object],
) -> bytes:
    """Build a public, deterministic XLSX report without Telegram identifiers.

    Only explicitly declared public fields are read from duck-typed inputs. Extra
    attributes and mapping keys are ignored, so Telegram IDs, names and usernames
    cannot leak into cells or workbook metadata.
    """

    clean_metadata = _coerce_metadata(metadata)
    clean_rounds = tuple(_coerce_round(round_data) for round_data in rounds)
    if not clean_rounds:
        raise ValueError("Для отчёта нужен хотя бы один раунд жеребьёвки.")
    round_numbers = [round_data.draw_round for round_data in clean_rounds]
    if len(set(round_numbers)) != len(round_numbers):
        raise ValueError("Номера раундов жеребьёвки не должны повторяться.")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Итоги"
    _build_summary_sheet(summary, clean_metadata, clean_rounds)

    one_round = len(clean_rounds) == 1
    for round_data in clean_rounds:
        sheet_name = "Участники" if one_round else f"Раунд {round_data.draw_round}"
        worksheet = workbook.create_sheet(sheet_name)
        _build_round_sheet(worksheet, round_data)

    workbook.properties.creator = "Telegram Twitch Giveaway Bot"
    workbook.properties.title = "Публичный отчёт о розыгрыше"
    workbook.properties.subject = "Прозрачные результаты жеребьёвки"
    workbook.properties.description = "Публичный отчёт без персональных идентификаторов"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False

    output = BytesIO()
    try:
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()
        output.close()


def _coerce_metadata(source: object) -> GiveawayReportMetadata:
    min_minutes = _finite_number(_read(source, "min_minutes"), "Минимум минут")
    if min_minutes <= 0:
        raise ValueError("Минимум минут должен быть положительным числом.")
    eligible_raw = _read(source, "eligible_count", None)
    eligible_count = (
        None
        if eligible_raw is None
        else _integer(eligible_raw, "Количество допущенных", minimum=0)
    )
    return GiveawayReportMetadata(
        giveaway_id=_integer(_read(source, "giveaway_id"), "ID розыгрыша", minimum=1),
        title=_clean_text(_read(source, "title")),
        prize=_clean_text(_read(source, "prize", "")),
        min_minutes=min_minutes,
        min_messages=_integer(
            _read(source, "min_messages"), "Минимум сообщений", minimum=1
        ),
        winner_count=_integer(
            _read(source, "winner_count"), "Количество победителей", minimum=1
        ),
        min_participants=_integer(
            _read(source, "min_participants"),
            "Минимум допущенных участников",
            minimum=1,
        ),
        message_interval_seconds=_integer(
            _read(source, "message_interval_seconds"),
            "Интервал сообщений",
            minimum=0,
        ),
        started_at=_read(source, "started_at", None),
        end_at=_read(source, "end_at", None),
        finished_at=_read(source, "finished_at", None),
        eligible_count=eligible_count,
    )


def _coerce_round(source: object) -> GiveawayReportRound:
    participants_raw = _read(source, "participants")
    if isinstance(participants_raw, (str, bytes, bytearray)):
        raise ValueError("Участники раунда должны быть последовательностью записей.")
    try:
        participants = tuple(
            _coerce_participant(participant) for participant in participants_raw
        )
    except TypeError as error:
        raise ValueError("Участники раунда должны быть последовательностью записей.") from error
    return GiveawayReportRound(
        draw_round=_integer(
            _read(source, "draw_round"), "Номер раунда", minimum=1
        ),
        forced=_boolean(_read(source, "forced"), "Принудительный режим"),
        participants=participants,
    )


def _coerce_participant(source: object) -> GiveawayReportParticipant:
    admitted = _boolean(_read(source, "admitted"), "Признак допуска")
    result = _clean_text(_read(source, "result"))
    if result not in ALLOWED_RESULTS:
        raise ValueError(
            "Результат должен быть «Победитель», «Не выбран» или «Не допущен»."
        )
    if admitted == (result == RESULT_NOT_ELIGIBLE):
        raise ValueError("Признак допуска не соответствует результату участника.")

    random_raw = _read(source, "random_number", None)
    random_number: str | None
    if random_raw is None or str(random_raw).strip() == "":
        random_number = None
    else:
        random_number = _clean_text(random_raw).strip()
        if _DECIMAL_IDENTIFIER.fullmatch(random_number) is None:
            raise ValueError("Случайное число должно быть десятичным идентификатором.")
        if int(random_number) > 2**63 - 1:
            raise ValueError("Случайное число должно помещаться в 63 бита.")

    place_raw = _read(source, "eligible_place", None)
    eligible_place = (
        None
        if place_raw is None
        else _integer(place_raw, "Место среди допущенных", minimum=1)
    )
    minutes = _finite_number(_read(source, "minutes"), "Количество минут")
    if minutes < 0:
        raise ValueError("Количество минут не может быть отрицательным.")
    return GiveawayReportParticipant(
        twitch_login=_clean_text(_read(source, "twitch_login")),
        minutes=minutes,
        messages=_integer(
            _read(source, "messages"), "Количество сообщений", minimum=0
        ),
        admitted=admitted,
        random_number=random_number,
        eligible_place=eligible_place,
        result=result,
    )


def _build_summary_sheet(
    worksheet: Worksheet,
    metadata: GiveawayReportMetadata,
    rounds: Sequence[GiveawayReportRound],
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _NAVY
    worksheet.freeze_panes = "A4"
    worksheet.merge_cells("A1:E1")
    _set_text(worksheet["A1"], "Публичный отчёт о розыгрыше")
    worksheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=_WHITE)
    worksheet["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 32

    worksheet.merge_cells("A2:E2")
    _set_text(
        worksheet["A2"],
        "Отчёт содержит только публичные Twitch-данные и результаты жеребьёвки.",
    )
    worksheet["A2"].font = Font(name="Aptos", size=10, italic=True, color=_MUTED)
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[2].height = 28

    row = 4
    row = _summary_section(worksheet, row, "Розыгрыш")
    row = _summary_value(worksheet, row, "ID розыгрыша", metadata.giveaway_id)
    row = _summary_value(worksheet, row, "Название", metadata.title)
    row = _summary_value(
        worksheet, row, "Награда", metadata.prize or "Не указана"
    )
    row = _summary_value(worksheet, row, "Начало", metadata.started_at, date_value=True)
    row = _summary_value(
        worksheet,
        row,
        "Плановое завершение",
        metadata.end_at if metadata.end_at is not None else "Не указано",
        date_value=metadata.end_at is not None,
    )
    row = _summary_value(
        worksheet,
        row,
        "Фактическое завершение",
        metadata.finished_at if metadata.finished_at is not None else "Не указано",
        date_value=metadata.finished_at is not None,
    )

    row += 1
    row = _summary_section(worksheet, row, "Условия")
    row = _summary_value(worksheet, row, "Минимум времени, мин", metadata.min_minutes)
    row = _summary_value(worksheet, row, "Минимум сообщений", metadata.min_messages)
    row = _summary_value(
        worksheet,
        row,
        "Интервал между сообщениями, сек",
        metadata.message_interval_seconds,
    )
    row = _summary_value(worksheet, row, "Победителей", metadata.winner_count)
    row = _summary_value(
        worksheet,
        row,
        "Минимум допущенных участников",
        metadata.min_participants,
    )
    if metadata.eligible_count is not None:
        row = _summary_value(
            worksheet,
            row,
            "Допущено при завершении",
            metadata.eligible_count,
        )

    row += 1
    row = _summary_section(worksheet, row, "История жеребьёвки")
    headers = ("Раунд", "Принудительный", "Участников", "Допущено", "Победителей")
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=column)
        _set_text(cell, header)
        _style_table_header(cell)
    first_data_row = row + 1
    for round_data in rounds:
        admitted_count = sum(person.admitted for person in round_data.participants)
        winner_count = sum(
            person.result == RESULT_WINNER for person in round_data.participants
        )
        values: tuple[object, ...] = (
            round_data.draw_round,
            "Да" if round_data.forced else "Нет",
            len(round_data.participants),
            admitted_count,
            winner_count,
        )
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row + 1, column=column)
            if isinstance(value, str):
                _set_text(cell, value)
            else:
                cell.value = value
                cell.number_format = "#,##0"
            cell.font = Font(name="Aptos", size=10, color=_TEXT)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if column != 1 else "right",
                vertical="center",
            )
        row += 1
    table = Table(displayName="RoundsSummary", ref=f"A{first_data_row - 1}:E{row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    widths = {"A": 36, "B": 30, "C": 18, "D": 18, "E": 18}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.auto_filter.ref = f"A{first_data_row - 1}:E{row}"
    worksheet.print_area = f"A1:E{row}"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.3
    worksheet.page_margins.right = 0.3


def _build_round_sheet(worksheet: Worksheet, round_data: GiveawayReportRound) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = _BLUE
    worksheet.freeze_panes = "A5"
    worksheet.merge_cells("A1:H1")
    _set_text(worksheet["A1"], f"Результаты жеребьёвки — раунд {round_data.draw_round}")
    worksheet["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=_WHITE)
    worksheet["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30

    admitted_count = sum(person.admitted for person in round_data.participants)
    winner_count = sum(
        person.result == RESULT_WINNER for person in round_data.participants
    )
    worksheet.merge_cells("A2:H2")
    _set_text(
        worksheet["A2"],
        (
            f"Участников: {len(round_data.participants)}  •  "
            f"Допущено: {admitted_count}  •  Победителей: {winner_count}  •  "
            f"Принудительный режим: {'Да' if round_data.forced else 'Нет'}"
        ),
    )
    worksheet["A2"].font = Font(name="Aptos", size=10, color=_MUTED)
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[2].height = 25

    headers = (
        "№",
        "Twitch",
        "Минуты",
        "Сообщения",
        "Допущен",
        "Случайное число",
        "Место среди допущенных",
        "Результат",
    )
    header_row = 4
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column)
        _set_text(cell, header)
        _style_table_header(cell)
    worksheet.row_dimensions[header_row].height = 30

    for index, participant in enumerate(round_data.participants, start=1):
        row = header_row + index
        values: tuple[object | None, ...] = (
            index,
            participant.twitch_login,
            participant.minutes,
            participant.messages,
            "Да" if participant.admitted else "Нет",
            participant.random_number,
            participant.eligible_place,
            participant.result,
        )
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column)
            if column in {2, 5, 6, 8}:
                _set_text(cell, "" if value is None else value)
            else:
                cell.value = value
            cell.font = Font(name="Aptos", size=10, color=_TEXT)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if column in {2, 8} else "center",
                vertical="center",
            )
        worksheet.cell(row=row, column=1).number_format = "#,##0"
        worksheet.cell(row=row, column=3).number_format = "0.00"
        worksheet.cell(row=row, column=4).number_format = "#,##0"
        score_cell = worksheet.cell(row=row, column=6)
        if participant.random_number is None:
            score_cell.value = None
        elif int(participant.random_number) <= _EXACT_NUMERIC_SCORE_MAX:
            # New draws use a 12-digit range that Excel can represent exactly.
            # Storing it as a number avoids scientific notation in viewers while
            # the custom format preserves leading zeroes.
            score_cell.value = int(participant.random_number)
            score_cell.number_format = "000000000000"
        else:
            # Backward compatibility for an XLSX made from an older saved draw
            # whose score exceeded Excel's exact numeric range.
            _set_text(score_cell, participant.random_number)
            score_cell.number_format = "@"
            score_cell.quotePrefix = True
        worksheet.cell(row=row, column=7).number_format = "#,##0"
        _style_status_cells(worksheet, row, participant)
        worksheet.row_dimensions[row].height = 22

    last_row = max(header_row, header_row + len(round_data.participants))
    table = Table(
        displayName=f"Round{round_data.draw_round}Participants",
        ref=f"A{header_row}:H{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.auto_filter.ref = f"A{header_row}:H{last_row}"

    widths = {
        "A": 8,
        "B": 24,
        "C": 14,
        "D": 14,
        "E": 13,
        "F": 23,
        "G": 25,
        "H": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.print_title_rows = f"1:{header_row}"
    worksheet.print_area = f"A1:H{last_row}"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25


def _summary_section(worksheet: Worksheet, row: int, title: str) -> int:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = worksheet.cell(row=row, column=1)
    _set_text(cell, title)
    cell.fill = PatternFill("solid", fgColor=_PALE_BLUE)
    cell.font = Font(name="Aptos", size=11, bold=True, color=_NAVY)
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row].height = 24
    return row + 1


def _summary_value(
    worksheet: Worksheet,
    row: int,
    label: str,
    value: object,
    *,
    date_value: bool = False,
) -> int:
    label_cell = worksheet.cell(row=row, column=1)
    _set_text(label_cell, label)
    label_cell.font = Font(name="Aptos", size=10, bold=True, color=_TEXT)
    label_cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    label_cell.border = _THIN_BORDER
    label_cell.alignment = Alignment(vertical="center")

    worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    value_cell = worksheet.cell(row=row, column=2)
    if date_value:
        converted = _excel_datetime(value)
        if isinstance(converted, datetime):
            # This is a read-only summary value. A preformatted public string is
            # rendered consistently by Excel, LibreOffice and Telegram previews.
            _set_text(value_cell, converted.strftime("%d.%m.%Y %H:%M (МСК)"))
        else:
            _set_text(value_cell, converted)
    elif isinstance(value, bool):
        _set_text(value_cell, "Да" if value else "Нет")
    elif isinstance(value, (int, float)):
        value_cell.value = value
        value_cell.number_format = "0.00" if isinstance(value, float) else "#,##0"
    else:
        _set_text(value_cell, value)
    value_cell.font = Font(name="Aptos", size=10, color=_TEXT)
    value_cell.border = _THIN_BORDER
    value_cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[row].height = 22
    return row + 1


def _style_table_header(cell: Cell) -> None:
    cell.fill = PatternFill("solid", fgColor=_BLUE)
    cell.font = Font(name="Aptos", size=10, bold=True, color=_WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="medium", color=_NAVY))


def _style_status_cells(
    worksheet: Worksheet, row: int, participant: GiveawayReportParticipant
) -> None:
    admitted_cell = worksheet.cell(row=row, column=5)
    result_cell = worksheet.cell(row=row, column=8)
    if participant.admitted:
        admitted_cell.fill = PatternFill("solid", fgColor=_GREEN)
        admitted_cell.font = Font(name="Aptos", size=10, bold=True, color=_GREEN_TEXT)
    else:
        admitted_cell.fill = PatternFill("solid", fgColor=_RED)
        admitted_cell.font = Font(name="Aptos", size=10, bold=True, color=_RED_TEXT)

    if participant.result == RESULT_WINNER:
        result_cell.fill = PatternFill("solid", fgColor=_GREEN)
        result_cell.font = Font(name="Aptos", size=10, bold=True, color=_GREEN_TEXT)
    elif participant.result == RESULT_NOT_ELIGIBLE:
        result_cell.fill = PatternFill("solid", fgColor=_RED)
        result_cell.font = Font(name="Aptos", size=10, bold=True, color=_RED_TEXT)
    else:
        result_cell.fill = PatternFill("solid", fgColor=_YELLOW)
        result_cell.font = Font(name="Aptos", size=10, bold=True, color=_YELLOW_TEXT)


def _read(source: object, field: str, default: object = _MISSING) -> object:
    if isinstance(source, Mapping):
        if field in source:
            return source[field]
    elif hasattr(source, field):
        return getattr(source, field)
    if default is _MISSING:
        raise ValueError(f"Не указано обязательное поле отчёта: {field}.")
    return default


def _clean_text(value: object) -> str:
    text = _ILLEGAL_XML_CHARACTERS.sub("", str(value))
    return text


def _set_text(cell: Cell, value: object) -> None:
    text = _clean_text(value)
    cell.value = text
    cell.data_type = "s"
    if text.startswith(_FORMULA_PREFIXES):
        cell.quotePrefix = True


def _integer(
    value: object,
    label: str,
    *,
    minimum: int | None = None,
    maximum: int | None = None,
) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label} должно быть целым числом.")
    try:
        result = int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label} должно быть целым числом.") from error
    if isinstance(value, float) and not value.is_integer():
        raise ValueError(f"{label} должно быть целым числом.")
    if minimum is not None and result < minimum:
        raise ValueError(f"{label} должно быть не меньше {minimum}.")
    if maximum is not None and result > maximum:
        raise ValueError(f"{label} должно быть не больше {maximum}.")
    return result


def _finite_number(value: object, label: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} должно быть числом.")
    try:
        result = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label} должно быть числом.") from error
    if not math.isfinite(result):
        raise ValueError(f"{label} должно быть конечным числом.")
    return result


def _boolean(value: object, label: str) -> bool:
    if not isinstance(value, bool):
        raise ValueError(f"{label} должен быть логическим значением.")
    return value


def _excel_datetime(value: object) -> datetime | str:
    if isinstance(value, bool):
        return _clean_text(value)
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(MOSCOW_TIMEZONE).replace(tzinfo=None)
        return value
    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(value, MOSCOW_TIMEZONE).replace(tzinfo=None)
    return _clean_text(value)
