from __future__ import annotations

import re
from dataclasses import replace
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
import tempfile
import unittest
from unittest.mock import patch

from aiogram.enums import ChatType
from aiogram.exceptions import TelegramForbiddenError
from aiogram.filters.command import CommandObject
from aiogram.methods import SendDocument, SendMessage
from openpyxl import load_workbook

from app.config import Settings
from app.main import (
    build_router,
    handle_stream_state_change,
    owner_bot_commands,
    parse_giveaway_end_at,
)
from app.storage import Giveaway, Storage
from app.twitch_chat import TwitchChatState


class FakeMessage:
    def __init__(
        self,
        user_id: int,
        *,
        full_name: str = "User",
        username: str | None = "user_tg",
        chat_type: ChatType = ChatType.PRIVATE,
        chat_id: int | None = None,
    ) -> None:
        self.from_user = SimpleNamespace(
            id=user_id,
            full_name=full_name,
            username=username,
        )
        self.chat = SimpleNamespace(type=chat_type, id=chat_id or user_id)
        self.answers: list[tuple[str, dict[str, object]]] = []

    async def answer(self, text: str, **kwargs: object) -> None:
        self.answers.append((text, kwargs))

    @property
    def last_text(self) -> str:
        return self.answers[-1][0]


class FakeBot:
    def __init__(self) -> None:
        self.sent: list[tuple[int, str, dict[str, object]]] = []
        self.documents: list[tuple[int, object, dict[str, object]]] = []
        self.member_statuses: dict[int, str] = {}
        self.failed_send_ids: set[int] = set()
        self.failed_document_ids: set[int] = set()
        self.failed_member_ids: set[int] = set()

    async def get_me(self) -> SimpleNamespace:
        return SimpleNamespace(username="deadlock_otp_bot")

    async def send_message(self, chat_id: int, text: str, **kwargs: object) -> None:
        if chat_id in self.failed_send_ids:
            raise TelegramForbiddenError(
                method=SendMessage(chat_id=chat_id, text=text),
                message="bot was blocked by the user",
            )
        self.sent.append((chat_id, text, kwargs))

    async def send_document(
        self, chat_id: int, document: object, **kwargs: object
    ) -> None:
        if chat_id in self.failed_document_ids:
            raise TelegramForbiddenError(
                method=SendDocument(chat_id=chat_id, document=document),
                message="bot was blocked by the user",
            )
        self.documents.append((chat_id, document, kwargs))

    async def get_chat_member(self, _chat_id: int, user_id: int) -> SimpleNamespace:
        if user_id in self.failed_member_ids:
            raise TelegramForbiddenError(
                method=SendMessage(chat_id=user_id, text="membership check"),
                message="membership unavailable",
            )
        status = self.member_statuses.get(user_id, "member")
        return SimpleNamespace(status=status, is_member=status == "member")


class BotCommandTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.storage = Storage(Path(self.temporary_directory.name) / "bot.sqlite3")
        await self.storage.connect()
        self.settings = Settings(
            telegram_bot_token="999999:token",
            telegram_required_chat_id=-100222,
            telegram_channel_url="https://t.me/deadlock_otp",
            owner_telegram_id=1,
            twitch_channel="deadlock_otp",
            twitch_bot_login="deadlock_otp",
            twitch_oauth_token="token",
            twitch_excluded_logins=(),
            telegram_excluded_usernames=(),
            twitch_live_check_interval_seconds=10,
            giveaway_xlsx_enabled=True,
            database_path=self.storage.path,
        )
        self.bot = FakeBot()
        self.twitch_state = TwitchChatState()
        self.twitch_state.mark_connected()
        self.refresh_calls = 0
        self.stream_live_on_refresh = False
        self.viewer_fetch_calls = 0
        self.viewer_logins = ["alice_tv", "bob_tv"]
        self.twitch_announce_config_calls: list[tuple[bool, int | None]] = []
        self.twitch_chat_send_validation_calls = 0
        self.twitch_chat_send_validation_error: str | None = None

        async def refresh_stream_status() -> None:
            self.refresh_calls += 1
            if self.stream_live_on_refresh:
                self.twitch_state.mark_stream_live("Test stream", "now")
            else:
                self.twitch_state.mark_stream_offline()

        async def fetch_viewers() -> list[str]:
            self.viewer_fetch_calls += 1
            return list(self.viewer_logins)

        async def configure_twitch_announcements(
            enabled: bool, interval_minutes: int | None
        ) -> Giveaway | None:
            self.twitch_announce_config_calls.append((enabled, interval_minutes))
            return await self.storage.configure_twitch_announcements(
                enabled=enabled, interval_minutes=interval_minutes
            )

        async def validate_twitch_chat_send() -> None:
            self.twitch_chat_send_validation_calls += 1
            if self.twitch_chat_send_validation_error is not None:
                raise RuntimeError(self.twitch_chat_send_validation_error)

        router = build_router(
            self.settings,
            self.storage,
            self.bot,  # type: ignore[arg-type]
            self.twitch_state,
            refresh_stream_status=refresh_stream_status,
            fetch_viewers=fetch_viewers,
            configure_twitch_announcements=configure_twitch_announcements,
            validate_twitch_chat_send=validate_twitch_chat_send,
        )
        self.handlers = {
            handler.callback.__name__: handler.callback for handler in router.message.handlers
        }

    async def asyncTearDown(self) -> None:
        await self.storage.close()
        self.temporary_directory.cleanup()

    def message(
        self,
        user_id: int = 1,
        *,
        full_name: str = "Owner",
        username: str | None = "owner_tg",
    ) -> FakeMessage:
        return FakeMessage(user_id, full_name=full_name, username=username)

    async def giveaway(self, message: FakeMessage, args: str) -> None:
        await self.handlers["giveaway_command"](
            message, CommandObject(command="giveaway", args=args)
        )

    async def shortcut(
        self, message: FakeMessage, command: str, args: str | None = None
    ) -> None:
        await self.handlers["giveaway_shortcut"](
            message, CommandObject(command=command, args=args)
        )

    async def confirm_finish(self, message: FakeMessage) -> None:
        await self.giveaway(message, "finish")
        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("Подтвердите завершение", message.last_text)
        await self.giveaway(message, "finish confirm")

    async def register(
        self,
        giveaway: Giveaway,
        telegram_id: int,
        telegram_name: str,
        telegram_username: str,
        twitch_login: str,
        twitch_id: str,
    ) -> None:
        code = await self.storage.create_link_code(
            telegram_id,
            telegram_name,
            giveaway.id,
            telegram_username,
        )
        status, linked_id, _ = await self.storage.claim_link_code(
            code, twitch_login, twitch_id
        )
        self.assertEqual((status, linked_id), ("linked", telegram_id))

    async def set_activity(
        self, giveaway_id: int, twitch_login: str, *, seconds: int, messages: int
    ) -> None:
        await self.storage._db.execute(
            """INSERT INTO giveaway_activity
               (giveaway_id, twitch_login, seconds, messages, presence_started_at)
               VALUES (?, ?, ?, ?, NULL)
               ON CONFLICT(giveaway_id, twitch_login) DO UPDATE SET
                   seconds = excluded.seconds,
                   messages = excluded.messages,
                   presence_started_at = NULL""",
            (giveaway_id, twitch_login, seconds, messages),
        )
        await self.storage._db.commit()

    def report_workbook(self, index: int = -1):
        document = self.bot.documents[index][1]
        data = getattr(document, "data")
        return load_workbook(BytesIO(data), data_only=False)

    async def test_start_command_explains_per_giveaway_registration(self) -> None:
        message = self.message(101, full_name="Alice", username="alice_tg")

        await self.handlers["start"](
            message, CommandObject(command="start", args=None)
        )

        self.assertIn("/link", message.last_text)
        self.assertIn("заново для каждого розыгрыша", message.last_text)

    async def test_link_requires_active_giveaway_and_issues_scoped_code(self) -> None:
        message = self.message(101, full_name="Alice", username="alice_tg")
        with patch("app.main.time.time", return_value=1000):
            await self.handlers["link"](message)
        self.assertIn("нет активного розыгрыша", message.last_text)

        giveaway = await self.storage.start_giveaway(10, 2, title="Тест")
        message.answers.clear()
        with patch("app.main.time.time", return_value=1011):
            await self.handlers["link"](message)

        match = re.search(r"!link ([A-Z0-9]{8})", message.last_text)
        self.assertIsNotNone(match)
        first_code = match.group(1)
        message.answers.clear()
        with patch("app.main.time.time", return_value=1022):
            await self.handlers["link"](message)
        repeated_match = re.search(r"!link ([A-Z0-9]{8})", message.last_text)
        self.assertIsNotNone(repeated_match)
        self.assertEqual(repeated_match.group(1), first_code)
        row = await (
            await self.storage._db.execute(
                "SELECT giveaway_id, telegram_username FROM link_codes WHERE code = ?",
                (first_code,),
            )
        ).fetchone()
        self.assertEqual((row["giveaway_id"], row["telegram_username"]), (giveaway.id, "alice_tg"))

    async def test_link_can_be_completed_before_stream_starts(self) -> None:
        giveaway = await self.storage.start_giveaway(10, 2, title="До стрима")
        self.twitch_state.mark_stream_offline()
        message = self.message(101, full_name="Alice", username="alice_tg")

        await self.handlers["link"](message)

        match = re.search(r"!link ([A-Z0-9]{8})", message.last_text)
        self.assertIsNotNone(match)
        status, telegram_id, _ = await self.storage.claim_link_code(
            match.group(1), "alice_tv", "777"
        )
        stats = await self.storage.participant_stats(giveaway, 101)
        self.assertEqual((status, telegram_id), ("linked", 101))
        self.assertIsNotNone(stats)
        self.assertEqual((stats.seconds, stats.messages), (0, 0))

    async def test_link_after_registration_reports_existing_twitch_account(self) -> None:
        giveaway = await self.storage.start_giveaway(10, 2, title="Повторная связь")
        message = self.message(101, full_name="Alice", username="alice_tg")

        with patch("app.main.time.time", return_value=1000):
            await self.handlers["link"](message)
            match = re.search(r"!link ([A-Z0-9]{8})", message.last_text)
            self.assertIsNotNone(match)
            await self.storage.claim_link_code(match.group(1), "alice_tv", "777")

        with patch("app.main.time.time", return_value=1001):
            await self.handlers["link"](message)

        self.assertIn("уже зарегистрированы", message.last_text)
        self.assertIn("Повторно выполнять /link не нужно", message.last_text)
        self.assertIn(
            '<a href="https://www.twitch.tv/alice_tv">alice_tv</a>',
            message.last_text,
        )
        self.assertNotIn("Слишком часто", message.last_text)
        row = await (
            await self.storage._db.execute(
                "SELECT COUNT(*) AS total FROM link_codes WHERE telegram_user_id = 101"
            )
        ).fetchone()
        self.assertEqual(row["total"], 0)

    async def test_public_link_command_has_cooldown(self) -> None:
        await self.storage.start_giveaway(10, 2, title="Cooldown")
        message = self.message(101, full_name="Alice", username="alice_tg")

        with patch("app.main.time.time", return_value=1000):
            await self.handlers["link"](message)
        with patch("app.main.time.time", return_value=1005):
            await self.handlers["link"](message)

        self.assertIn("Слишком часто", message.last_text)
        self.assertIn("через 5 сек", message.last_text)
        row = await (
            await self.storage._db.execute(
                "SELECT COUNT(*) AS total FROM link_codes WHERE telegram_user_id = 101"
            )
        ).fetchone()
        self.assertEqual(row["total"], 1)

    async def test_link_waits_until_twitch_chat_collector_is_connected(self) -> None:
        await self.storage.start_giveaway(10, 2, title="Тест")
        self.twitch_state.mark_disconnected()
        message = self.message(101, full_name="Alice", username="alice_tg")

        await self.handlers["link"](message)

        self.assertIn("переподключается", message.last_text)
        row = await (
            await self.storage._db.execute(
                "SELECT COUNT(*) AS total FROM link_codes WHERE telegram_user_id = 101"
            )
        ).fetchone()
        self.assertEqual(row["total"], 0)

    async def test_announcement_deep_link_start_payload_issues_link_code(self) -> None:
        giveaway = await self.storage.start_giveaway(1, 1, title="Deep link")
        message = self.message(101, full_name="Alice", username="alice_tg")

        await self.handlers["start"](
            message, CommandObject(command="start", args="link")
        )

        self.assertIn("!link", message.last_text)
        self.assertIn(
            '<a href="https://www.twitch.tv/deadlock_otp">deadlock_otp</a>',
            message.last_text,
        )
        row = await (
            await self.storage._db.execute(
                "SELECT giveaway_id FROM link_codes WHERE telegram_user_id = 101"
            )
        ).fetchone()
        self.assertEqual(row["giveaway_id"], giveaway.id)

    async def test_deep_link_start_uses_link_cooldown(self) -> None:
        await self.storage.start_giveaway(10, 2, title="Deep cooldown")
        message = self.message(101, full_name="Alice", username="alice_tg")

        with patch("app.main.time.time", return_value=1000):
            await self.handlers["start"](
                message, CommandObject(command="start", args="link")
            )
        with patch("app.main.time.time", return_value=1001):
            await self.handlers["start"](
                message, CommandObject(command="start", args="link")
            )

        self.assertIn("Слишком часто", message.last_text)
        self.assertIn("через 9 сек", message.last_text)

    async def test_status_refreshes_stream_and_shows_personal_progress(self) -> None:
        giveaway = await self.storage.start_giveaway(10, 2, title="Тест")
        await self.register(giveaway, 101, "Alice", "alice_tg", "alice_tv", "777")
        await self.set_activity(giveaway.id, "alice_tv", seconds=300, messages=1)
        message = self.message(101, full_name="Alice", username="alice_tg")

        await self.handlers["tracking_status"](message)

        self.assertEqual(self.refresh_calls, 1)
        self.assertIn("стрим не live", message.last_text)
        self.assertIn("Twitch: <code>alice_tv</code>", message.last_text)
        self.assertIn("Время: 5 мин из 10 мин", message.last_text)
        self.assertIn("Сообщения: 1 из 2", message.last_text)
        self.assertIn("Зарегистрировано участников: <b>1</b>", message.last_text)
        self.assertNotIn("Диагностика для владельца", message.last_text)

    async def test_public_status_command_has_cooldown(self) -> None:
        message = self.message(101, full_name="Alice", username="alice_tg")

        with patch("app.main.time.time", return_value=1000):
            await self.handlers["tracking_status"](message)
        with patch("app.main.time.time", return_value=1002):
            await self.handlers["tracking_status"](message)

        self.assertEqual(self.refresh_calls, 1)
        self.assertIn("Слишком часто", message.last_text)
        self.assertIn("через 3 сек", message.last_text)

    async def test_owner_public_commands_are_not_rate_limited(self) -> None:
        owner = self.message(1)

        with patch("app.main.time.time", return_value=1000):
            await self.handlers["tracking_status"](owner)
        with patch("app.main.time.time", return_value=1001):
            await self.handlers["tracking_status"](owner)

        self.assertEqual(self.refresh_calls, 2)
        self.assertNotIn("Слишком часто", owner.last_text)

    async def test_owner_status_contains_diagnostics_but_public_status_does_not(self) -> None:
        await self.storage.start_giveaway(1, 1)
        owner = self.message(1)
        user = self.message(101)

        await self.handlers["tracking_status"](owner)
        await self.handlers["tracking_status"](user)

        self.assertIn("Диагностика для владельца", owner.last_text)
        self.assertNotIn("Диагностика для владельца", user.last_text)

    async def test_viewers_is_available_only_to_owner(self) -> None:
        owner = self.message(1)
        user = self.message(101)
        self.stream_live_on_refresh = True

        await self.handlers["viewers_command"](user)
        self.assertEqual(user.answers, [])
        self.assertEqual(self.viewer_fetch_calls, 0)

        await self.handlers["viewers_command"](owner)
        self.assertEqual(self.viewer_fetch_calls, 1)
        combined = "\n".join(answer[0] for answer in owner.answers)
        self.assertIn("Участники Twitch-чата сейчас:</b> 2", combined)
        self.assertIn("@alice_tv", combined)
        self.assertIn("@bob_tv", combined)

    async def test_viewers_reports_offline_without_requesting_chatters(self) -> None:
        owner = self.message(1)

        await self.handlers["viewers_command"](owner)

        self.assertEqual(self.viewer_fetch_calls, 0)
        self.assertIn("offline", owner.last_text)

    async def test_stream_announcement_respects_offline_cooldown_and_stream_id(self) -> None:
        self.twitch_state.mark_stream_live("Первый стрим", "stream-1")
        with patch("app.main.time.time", return_value=100):
            await handle_stream_state_change(
                True, self.bot, self.settings, self.storage, self.twitch_state  # type: ignore[arg-type]
            )

        self.assertEqual(len(self.bot.sent), 1)
        self.assertEqual(self.bot.sent[0][0], self.settings.telegram_required_chat_id)
        self.assertIn("🔴 <b>Стрим начался!</b>", self.bot.sent[0][1])
        self.assertIn("Первый стрим", self.bot.sent[0][1])
        self.assertIn("https://www.twitch.tv/deadlock_otp", self.bot.sent[0][1])

        with patch("app.main.time.time", return_value=200):
            await handle_stream_state_change(
                True, self.bot, self.settings, self.storage, self.twitch_state  # type: ignore[arg-type]
            )
        self.assertEqual(len(self.bot.sent), 1)

        with patch("app.main.time.time", return_value=1_000):
            await handle_stream_state_change(
                False, self.bot, self.settings, self.storage, self.twitch_state  # type: ignore[arg-type]
            )
        self.twitch_state.mark_stream_live("Второй стрим", "stream-2")
        with patch("app.main.time.time", return_value=2_799):
            await handle_stream_state_change(
                True, self.bot, self.settings, self.storage, self.twitch_state  # type: ignore[arg-type]
            )
        self.assertEqual(len(self.bot.sent), 1)

        with patch("app.main.time.time", return_value=3_000):
            await handle_stream_state_change(
                False, self.bot, self.settings, self.storage, self.twitch_state  # type: ignore[arg-type]
            )
        self.twitch_state.mark_stream_live("Третий стрим", "stream-3")
        with patch("app.main.time.time", return_value=4_800):
            await handle_stream_state_change(
                True, self.bot, self.settings, self.storage, self.twitch_state  # type: ignore[arg-type]
            )

        self.assertEqual(len(self.bot.sent), 2)
        self.assertIn("Третий стрим", self.bot.sent[-1][1])
        self.assertEqual(
            await self.storage.runtime_state("last_announced_stream_started_at"),
            "stream-3",
        )

    async def test_admin_create_announce_status_and_participants(self) -> None:
        owner = self.message(1)
        await self.giveaway(
            owner,
            "start 10 2 2 30 3 --end 31.12.2099 23:59 Тестовый розыгрыш | Steam key",
        )
        giveaway = await self.storage.active_giveaway()
        self.assertIsNotNone(giveaway)
        self.assertEqual(
            (
                giveaway.title,
                giveaway.prize,
                giveaway.winner_count,
                giveaway.message_interval_seconds,
                giveaway.min_participants,
                giveaway.end_at is not None,
            ),
            ("Тестовый розыгрыш", "Steam key", 2, 30, 3, True),
        )
        self.assertIn("ещё не анонсирован", owner.last_text)
        self.assertEqual(self.bot.sent, [])

        await self.giveaway(owner, "announce_start")
        self.assertIn("https://t.me/deadlock_otp_bot?start=link", self.bot.sent[-1][1])
        self.assertIn("Steam key", self.bot.sent[-1][1])
        self.assertIn("31.12.2099 в 23:59 (МСК)", self.bot.sent[-1][1])

        await self.register(giveaway, 101, "Alice", "alice_tg", "alice_tv", "777")
        await self.set_activity(giveaway.id, "alice_tv", seconds=600, messages=2)
        await self.giveaway(owner, "status")
        self.assertIn("Зарегистрировались в розыгрыше: 1", owner.last_text)
        self.assertIn("Уже выполнили технические условия: 1", owner.last_text)

        await self.giveaway(owner, "participants Тестовый розыгрыш")
        self.assertIn("alice_tv", owner.last_text)
        self.assertIn("@alice_tg", owner.last_text)

    async def test_owner_can_edit_every_active_giveaway_parameter(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 0 1 Исходный | Старый приз")
        original = await self.storage.active_giveaway()
        self.assertIsNotNone(original)

        edits = (
            ("minutes 100", "min_seconds", 6000),
            ("messages 12", "min_messages", 12),
            ("winners 2", "winner_count", 2),
            ("interval 45", "message_interval_seconds", 45),
            ("participants 7", "min_participants", 7),
            (
                "end 31.12.2099 23:59",
                "end_at",
                parse_giveaway_end_at("31.12.2099", "23:59"),
            ),
            ("title <Первый & главный>", "title", "<Первый & главный>"),
            ("prize Steam 2000 ₽ | регион РФ", "prize", "Steam 2000 ₽ | регион РФ"),
        )
        for edit_args, attribute, expected in edits:
            with self.subTest(edit_args=edit_args):
                await self.giveaway(owner, f"edit {edit_args}")
                updated = await self.storage.active_giveaway()
                self.assertIsNotNone(updated)
                self.assertEqual(getattr(updated, attribute), expected)
                self.assertIn("Параметр активного розыгрыша обновлён", owner.last_text)

        updated = await self.storage.active_giveaway()
        self.assertEqual(updated.id, original.id)
        self.assertEqual(updated.started_at, original.started_at)
        self.assertIsNotNone(
            await self.storage.latest_giveaway_by_title("<первый & ГЛАВНЫЙ>")
        )
        self.assertIsNone(await self.storage.latest_giveaway_by_title("Исходный"))

        await self.giveaway(owner, "status")
        self.assertIn("&lt;Первый &amp; главный&gt;", owner.last_text)
        self.assertIn("Условия: 100 мин, 12 сообщений", owner.last_text)
        self.assertIn("Интервал сообщений: 45 сек", owner.last_text)
        self.assertIn("Победителей: 2", owner.last_text)
        self.assertIn("не менее 7", owner.last_text)
        self.assertIn("Steam 2000 ₽ | регион РФ", owner.last_text)

        await self.giveaway(owner, "announce_start")
        announcement = self.bot.sent[-1][1]
        self.assertIn("Время просмотра трансляции: не менее <b>100 мин</b>", announcement)
        self.assertIn("Сообщения в Twitch-чате: не менее <b>12</b>", announcement)
        self.assertIn("Steam 2000 ₽ | регион РФ", announcement)
        self.assertIn("31.12.2099 в 23:59 (МСК)", announcement)

    async def test_edit_can_clear_prize_and_planned_end(self) -> None:
        owner = self.message(1)
        await self.giveaway(
            owner,
            "start 10 2 1 30 1 --end 31.12.2099 23:59 Очистка | Steam key",
        )

        await self.giveaway(owner, "edit prize clear")
        self.assertIn("не указана", owner.last_text)
        await self.giveaway(owner, "edit end clear")
        self.assertIn("не задано", owner.last_text)

        updated = await self.storage.active_giveaway()
        self.assertEqual(updated.prize, "")
        self.assertIsNone(updated.end_at)
        await self.giveaway(owner, "announce_start")
        self.assertNotIn("🎁 <b>Награда</b>", self.bot.sent[-1][1])
        self.assertNotIn("Плановое завершение", self.bot.sent[-1][1])

    async def test_edit_rejects_invalid_values_without_partial_changes(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 30 1 Неизменный | Приз")
        original = await self.storage.active_giveaway()
        invalid_edits = (
            ("minutes 0", "не меньше 1 минуты"),
            ("messages 0", "не меньше 1"),
            ("winners 101", "от 1 до 100"),
            ("interval -1", "не может быть отрицательным"),
            ("participants 100001", "от 1 до 100000"),
            ("end 01.01.2000 00:00", "должны быть в будущем"),
            (f"title {'я' * 121}", "длиннее 120"),
            (f"prize {'п' * 301}", "длиннее 300"),
            ("unknown 5", "Неизвестный параметр"),
            ("minutes много", "целым числом"),
            (f"minutes {'9' * 100}", "слишком большое количество минут"),
            (f"messages {'9' * 100}", "слишком большое количество сообщений"),
            (f"interval {'9' * 100}", "слишком большой интервал"),
        )

        for edit_args, expected_error in invalid_edits:
            with self.subTest(edit_args=edit_args):
                await self.giveaway(owner, f"edit {edit_args}")
                self.assertIn(expected_error, owner.last_text)
                self.assertEqual(await self.storage.active_giveaway(), original)

    async def test_edit_cancels_pending_finish_confirmation(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 30 1 Новые условия")
        await self.giveaway(owner, "finish")

        await self.giveaway(owner, "edit minutes 20")

        self.assertIn("Предыдущее подтверждение завершения отменено", owner.last_text)
        await self.giveaway(owner, "finish confirm")
        self.assertIn("Нет действующего подтверждения", owner.last_text)
        self.assertIsNotNone(await self.storage.active_giveaway())

    async def test_finish_confirmation_detects_out_of_band_rule_change(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 30 1 Контроль версии")
        await self.giveaway(owner, "finish")

        await self.storage.update_active_giveaway(min_minutes=20)
        await self.giveaway(owner, "finish confirm")

        self.assertIn("изменились параметры розыгрыша", owner.last_text)
        self.assertIsNotNone(await self.storage.active_giveaway())

    async def test_edit_same_value_keeps_pending_finish_confirmation(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 30 1 Без изменений")
        await self.giveaway(owner, "finish")

        await self.giveaway(owner, "edit minutes 10")
        self.assertIn("уже имеет указанное значение", owner.last_text)
        await self.giveaway(owner, "finish confirm")

        self.assertIsNone(await self.storage.active_giveaway())

    async def test_finish_uses_edited_thresholds_and_winner_count(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Новые правила")
        giveaway = await self.storage.active_giveaway()
        participants = (
            (101, "Alice", "alice_tv", "777", 120, 2),
            (102, "Bob", "bob_tv", "778", 120, 2),
            (103, "Eve", "eve_tv", "779", 60, 1),
        )
        for telegram_id, name, login, twitch_id, seconds, messages in participants:
            await self.register(
                giveaway,
                telegram_id,
                name,
                f"{name.lower()}_tg",
                login,
                twitch_id,
            )
            await self.set_activity(
                giveaway.id, login, seconds=seconds, messages=messages
            )

        for edit_args in (
            "minutes 2",
            "messages 2",
            "winners 2",
            "participants 2",
        ):
            await self.giveaway(owner, f"edit {edit_args}")
        await self.confirm_finish(owner)

        finished = await self.storage.latest_finished_giveaway()
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual(finished.eligible_count_at_finish, 2)
        self.assertEqual(finished.winner_count, 2)
        self.assertEqual(
            {winner.telegram_user_id for winner in winners},
            {101, 102},
        )

    async def test_finish_always_closes_and_announce_finish_works_without_winner(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 0 2 Без победителей")

        await self.giveaway(owner, "finish")

        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("Подтвердите завершение", owner.last_text)
        await self.giveaway(owner, "finish confirm")

        self.assertIsNone(await self.storage.active_giveaway())
        finished = await self.storage.latest_finished_giveaway()
        self.assertEqual(finished.eligible_count_at_finish, 0)
        self.assertIn("завершён без победителей", owner.last_text)
        self.assertEqual(self.bot.sent, [])

        await self.giveaway(owner, "announce_finish")
        self.assertIn("Розыгрыш завершён", self.bot.sent[-1][1])
        self.assertIn("Победители не выбраны", self.bot.sent[-1][1])
        self.assertIn("Допущенных участников: <b>0</b>", self.bot.sent[-1][1])

    async def test_force_finish_selects_winner_below_required_participant_count(self) -> None:
        owner = self.message(1)
        await self.giveaway(
            owner, "start 1 1 1 0 3 Принудительное завершение | Steam key"
        )
        giveaway = await self.storage.active_giveaway()
        await self.register(
            giveaway, 101, "Alice", "alice_tg", "alice_tv", "777"
        )
        await self.set_activity(
            giveaway.id, "alice_tv", seconds=60, messages=1
        )

        await self.giveaway(owner, "finish")

        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("/giveaway_finish force confirm", owner.last_text)
        await self.giveaway(owner, "finish force confirm")

        self.assertIsNone(await self.storage.active_giveaway())
        finished = await self.storage.latest_finished_giveaway()
        self.assertEqual(finished.eligible_count_at_finish, 1)
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual([winner.twitch_login for winner in winners], ["alice_tv"])
        self.assertIn("завершён принудительно", owner.last_text)
        private_messages = [sent for sent in self.bot.sent if sent[0] == 101]
        self.assertEqual(len(private_messages), 1)

        await self.giveaway(owner, "announce_finish")
        self.assertIn("@alice_tg", self.bot.sent[-1][1])
        self.assertIn("принудительном режиме", self.bot.sent[-1][1])

    async def test_finish_selects_only_candidate_who_satisfies_every_condition(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 10 2 1 0 1 Проверка победителя | Steam key")
        giveaway = await self.storage.active_giveaway()
        people = (
            (101, "Alice", "alice_tg", "alice_tv", "777", 600, 2),
            (102, "Bob", "bob_tg", "bob_tv", "778", 599, 2),
            (103, "Cara", "cara_tg", "cara_tv", "779", 600, 1),
            (104, "Dan", "dan_tg", "dan_tv", "780", 600, 2),
        )
        for telegram_id, name, tg_username, login, twitch_id, seconds, messages in people:
            await self.register(
                giveaway, telegram_id, name, tg_username, login, twitch_id
            )
            await self.set_activity(
                giveaway.id, login, seconds=seconds, messages=messages
            )
        self.bot.member_statuses[104] = "left"

        await self.confirm_finish(owner)

        finished = await self.storage.latest_finished_giveaway()
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual([winner.twitch_login for winner in winners], ["alice_tv"])
        private_messages = [sent for sent in self.bot.sent if sent[0] == 101]
        self.assertEqual(len(private_messages), 1)
        self.assertIn("Вы победили в розыгрыше", private_messages[0][1])
        self.assertIn("Проверка победителя", private_messages[0][1])
        self.assertIn("Steam key", private_messages[0][1])
        self.assertIn("alice_tv", private_messages[0][1])
        await self.giveaway(owner, "announce_finish")
        self.assertIn("@alice_tg", self.bot.sent[-1][1])
        self.assertIn("alice_tv", self.bot.sent[-1][1])
        self.assertNotIn("bob_tv", self.bot.sent[-1][1])
        self.assertNotIn("cara_tv", self.bot.sent[-1][1])
        self.assertNotIn("dan_tv", self.bot.sent[-1][1])

    async def test_finish_rolls_all_registrations_and_publishes_private_xlsx(self) -> None:
        owner = self.message(1)
        await self.giveaway(
            owner,
            "start 10 2 1 0 1 Прозрачный ролл | Пополнение Steam",
        )
        giveaway = await self.storage.active_giveaway()
        people = (
            (101, "Private Alice", "private_alice_tg", "alice_tv", "777", 600, 2),
            (102, "Private Bob", "private_bob_tg", "bob_tv", "778", 600, 2),
            (987654321012345, "Private Cara", "private_cara_tg", "cara_tv", "779", 599, 2),
        )
        for telegram_id, name, tg_username, login, twitch_id, seconds, messages in people:
            await self.register(
                giveaway, telegram_id, name, tg_username, login, twitch_id
            )
            await self.set_activity(
                giveaway.id, login, seconds=seconds, messages=messages
            )

        await self.giveaway(owner, "finish")
        self.assertEqual(self.bot.documents, [])
        with patch.object(Storage, "_unique_draw_scores", return_value=[100, 900, 500]):
            await self.giveaway(owner, "finish confirm")

        finished = await self.storage.latest_finished_giveaway()
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual([winner.twitch_login for winner in winners], ["bob_tv"])
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 1)
        self.assertEqual(len(self.bot.documents), 1)
        self.assertEqual(self.bot.documents[0][0], self.settings.owner_telegram_id)

        workbook = self.report_workbook()
        try:
            self.assertEqual(workbook.sheetnames, ["Итоги", "Участники"])
            worksheet = workbook["Участники"]
            rows = {
                worksheet.cell(row=row, column=2).value: {
                    "minutes": worksheet.cell(row=row, column=3).value,
                    "messages": worksheet.cell(row=row, column=4).value,
                    "admitted": worksheet.cell(row=row, column=5).value,
                    "score": worksheet.cell(row=row, column=6).value,
                    "place": worksheet.cell(row=row, column=7).value,
                    "result": worksheet.cell(row=row, column=8).value,
                }
                for row in range(5, 8)
            }
            self.assertEqual(set(rows), {"alice_tv", "bob_tv", "cara_tv"})
            self.assertEqual(rows["alice_tv"]["minutes"], 10)
            self.assertEqual(rows["bob_tv"]["messages"], 2)
            self.assertEqual(rows["bob_tv"]["score"], 900)
            self.assertEqual(rows["bob_tv"]["place"], 1)
            self.assertEqual(rows["bob_tv"]["result"], "Победитель")
            self.assertEqual(rows["alice_tv"]["result"], "Не выбран")
            self.assertEqual(rows["cara_tv"]["admitted"], "Нет")
            self.assertEqual(rows["cara_tv"]["result"], "Не допущен")
            all_values = {
                value
                for worksheet in workbook.worksheets
                for row in worksheet.iter_rows(values_only=True)
                for value in row
            }
            self.assertNotIn("Private Alice", all_values)
            self.assertNotIn("private_alice_tg", all_values)
            self.assertNotIn(987654321012345, all_values)
        finally:
            workbook.close()

        await self.giveaway(owner, "announce_finish")

        self.assertEqual(len(self.bot.documents), 2)
        self.assertEqual(
            self.bot.documents[-1][0], self.settings.telegram_required_chat_id
        )
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 1)
        public_workbook = self.report_workbook()
        try:
            public_scores = {
                public_workbook["Участники"].cell(row=row, column=2).value:
                public_workbook["Участники"].cell(row=row, column=6).value
                for row in range(5, 8)
            }
            self.assertEqual(
                public_scores,
                {login: values["score"] for login, values in rows.items()},
            )
        finally:
            public_workbook.close()

    async def test_disabled_xlsx_keeps_draw_data_but_sends_no_documents(self) -> None:
        disabled_settings = replace(self.settings, giveaway_xlsx_enabled=False)
        router = build_router(
            disabled_settings,
            self.storage,
            self.bot,  # type: ignore[arg-type]
            self.twitch_state,
        )
        handlers = {
            handler.callback.__name__: handler.callback for handler in router.message.handlers
        }
        owner = self.message(1)

        async def giveaway_command(args: str) -> None:
            await handlers["giveaway_command"](
                owner, CommandObject(command="giveaway", args=args)
            )

        await giveaway_command("start 1 1 1 0 1 Без таблицы")
        giveaway = await self.storage.active_giveaway()
        await self.register(
            giveaway, 101, "Alice", "alice_tg", "alice_tv", "777"
        )
        await self.set_activity(giveaway.id, "alice_tv", seconds=60, messages=1)

        await giveaway_command("finish")
        await giveaway_command("finish confirm")

        finished = await self.storage.latest_finished_giveaway()
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 1)
        self.assertEqual(
            [winner.twitch_login for winner in await self.storage.recorded_winners(finished)],
            ["alice_tv"],
        )
        self.assertEqual(self.bot.documents, [])
        private_text = "\n".join(text for text, _kwargs in owner.answers)
        self.assertIn("сохранены в базе данных", private_text)
        self.assertNotIn("XLSX", private_text)

        await giveaway_command("announce_finish")

        self.assertEqual(self.bot.documents, [])
        channel_messages = [
            text
            for chat_id, text, _kwargs in self.bot.sent
            if chat_id == self.settings.telegram_required_chat_id
        ]
        self.assertTrue(channel_messages)
        self.assertIn("alice_tv", "\n".join(channel_messages))
        sent_count = len(self.bot.sent)

        await giveaway_command("report")

        self.assertEqual(len(self.bot.documents), 1)
        self.assertEqual(self.bot.documents[0][0], disabled_settings.owner_telegram_id)
        self.assertEqual(len(self.bot.sent), sent_count)
        self.assertIn("XLSX-отчёт отправлен", owner.last_text)
        workbook = self.report_workbook()
        self.assertEqual(workbook.sheetnames, ["Итоги", "Участники"])

    async def test_membership_error_keeps_giveaway_and_progress_active(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Ошибка Telegram")
        giveaway = await self.storage.active_giveaway()
        await self.register(
            giveaway, 101, "Alice", "alice_tg", "alice_tv", "777"
        )
        await self.set_activity(giveaway.id, "alice_tv", seconds=60, messages=1)
        self.bot.failed_member_ids.add(101)

        await self.giveaway(owner, "finish")
        await self.giveaway(owner, "finish confirm")

        active = await self.storage.active_giveaway()
        self.assertEqual(active.id, giveaway.id)
        participants = await self.storage.giveaway_participants(active)
        self.assertEqual((participants[0].seconds, participants[0].messages), (60, 1))
        self.assertEqual(await self.storage.draw_rounds(giveaway), [])
        self.assertEqual(self.bot.documents, [])
        self.assertIn("Розыгрыш остаётся активным", owner.last_text)

    async def test_xlsx_failure_does_not_undo_finished_draw(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Ошибка XLSX")
        giveaway = await self.storage.active_giveaway()
        await self.register(
            giveaway, 101, "Alice", "alice_tg", "alice_tv", "777"
        )
        await self.set_activity(giveaway.id, "alice_tv", seconds=60, messages=1)
        self.bot.failed_document_ids.add(self.settings.owner_telegram_id)

        await self.confirm_finish(owner)

        self.assertIsNone(await self.storage.active_giveaway())
        finished = await self.storage.latest_finished_giveaway()
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 1)
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual([winner.telegram_user_id for winner in winners], [101])
        self.assertEqual(self.bot.documents, [])
        self.assertIn("XLSX не удалось", "\n".join(text for text, _ in owner.answers))

    async def test_reroll_creates_a_separate_report_round(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Два раунда")
        giveaway = await self.storage.active_giveaway()
        for telegram_id, name, login, twitch_id in (
            (101, "Alice", "alice_tv", "777"),
            (102, "Bob", "bob_tv", "778"),
        ):
            await self.register(
                giveaway,
                telegram_id,
                name,
                f"{name.lower()}_tg",
                login,
                twitch_id,
            )
            await self.set_activity(giveaway.id, login, seconds=60, messages=1)

        await self.giveaway(owner, "finish")
        with patch.object(Storage, "_unique_draw_scores", return_value=[900, 100]):
            await self.giveaway(owner, "finish confirm")
        with patch.object(Storage, "_unique_draw_scores", return_value=[800, 700]):
            await self.giveaway(owner, "reroll")

        finished = await self.storage.latest_finished_giveaway()
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 2)
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual(
            {winner.twitch_login for winner in winners}, {"alice_tv", "bob_tv"}
        )
        workbook = self.report_workbook()
        try:
            self.assertEqual(workbook.sheetnames, ["Итоги", "Раунд 1", "Раунд 2"])
        finally:
            workbook.close()

        await self.giveaway(owner, "announce_finish")
        self.assertEqual(self.bot.documents[-1][0], self.settings.telegram_required_chat_id)
        public_workbook = self.report_workbook()
        try:
            self.assertEqual(public_workbook.sheetnames, ["Итоги", "Раунд 1", "Раунд 2"])
        finally:
            public_workbook.close()

    async def test_reroll_cannot_bypass_minimum_after_finish_without_winner(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 2 Минимум не достигнут")
        giveaway = await self.storage.active_giveaway()
        await self.register(
            giveaway,
            101,
            "Alice",
            "alice_tg",
            "alice_tv",
            "701",
        )
        await self.set_activity(giveaway.id, "alice_tv", seconds=60, messages=1)

        await self.confirm_finish(owner)
        finished = await self.storage.latest_finished_giveaway()
        self.assertEqual(await self.storage.recorded_winners(finished), [])
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 1)
        document_count = len(self.bot.documents)

        await self.giveaway(owner, "reroll")

        self.assertIn("Подходящих участников не осталось", owner.last_text)
        self.assertEqual(len(await self.storage.draw_rounds(finished)), 1)
        self.assertEqual(await self.storage.recorded_winners(finished), [])
        self.assertEqual(len(self.bot.documents), document_count)

    async def test_owner_can_enable_configure_and_disable_twitch_announcements(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Twitch-анонсы")
        giveaway = await self.storage.active_giveaway()
        self.assertFalse(giveaway.twitch_announce_enabled)

        await self.shortcut(owner, "twitch_announce", "on 15")

        giveaway = await self.storage.active_giveaway()
        self.assertTrue(giveaway.twitch_announce_enabled)
        self.assertEqual(giveaway.twitch_announce_interval_seconds, 900)
        self.assertEqual(self.twitch_announce_config_calls[-1], (True, 15))
        self.assertEqual(self.twitch_chat_send_validation_calls, 1)
        self.assertIn("каждые <b>15 мин</b>", owner.last_text)
        self.assertIn("chat:edit", owner.last_text)

        await self.shortcut(owner, "twitch_announce", "status")
        self.assertIn("Twitch-анонсы розыгрыша: <b>включены</b>", owner.last_text)
        self.assertIn("стрим offline", owner.last_text)

        await self.giveaway(owner, "status")
        self.assertIn("Twitch-анонсы розыгрыша: <b>включены</b>", owner.last_text)

        await self.shortcut(owner, "twitch_announce", "off")
        giveaway = await self.storage.active_giveaway()
        self.assertFalse(giveaway.twitch_announce_enabled)
        self.assertEqual(self.twitch_announce_config_calls[-1], (False, None))

    async def test_twitch_announcement_interval_is_validated(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Интервал")

        await self.shortcut(owner, "twitch_announce", "on 0")

        self.assertIn("от 1 до 1440 минут", owner.last_text)
        giveaway = await self.storage.active_giveaway()
        self.assertFalse(giveaway.twitch_announce_enabled)

    async def test_twitch_announcements_are_not_enabled_without_chat_edit_scope(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Права Twitch")
        self.twitch_chat_send_validation_error = (
            "OAuth-токен Twitch не содержит право chat:edit"
        )

        await self.shortcut(owner, "twitch_announce", "on 15")

        self.assertIn("chat:edit", owner.last_text)
        giveaway = await self.storage.active_giveaway()
        self.assertFalse(giveaway.twitch_announce_enabled)
        self.assertEqual(self.twitch_announce_config_calls, [])

    async def test_finishing_giveaway_disables_twitch_announcements(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Завершение анонсов")
        await self.shortcut(owner, "twitch_announce", "on 15")

        await self.confirm_finish(owner)

        finished = await self.storage.latest_finished_giveaway()
        self.assertFalse(finished.twitch_announce_enabled)
        self.assertEqual(self.twitch_announce_config_calls[-1], (False, None))

    async def test_owner_can_participate_when_exclusion_fields_are_empty(self) -> None:
        owner = self.message(1, full_name="Owner", username="owner_tg")
        await self.giveaway(owner, "start 1 1 1 0 1 Участие владельца")
        giveaway = await self.storage.active_giveaway()
        await self.register(
            giveaway,
            1,
            "Owner",
            "owner_tg",
            "deadlock_otp",
            "channel-owner-id",
        )
        await self.set_activity(
            giveaway.id, "deadlock_otp", seconds=60, messages=1
        )

        await self.confirm_finish(owner)

        finished = await self.storage.latest_finished_giveaway()
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual(len(winners), 1)
        self.assertEqual(winners[0].telegram_user_id, 1)
        self.assertEqual(winners[0].twitch_login, "deadlock_otp")

    async def test_reroll_never_selects_recorded_winner_twice(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Перевыбор")
        giveaway = await self.storage.active_giveaway()
        for telegram_id, name, login, twitch_id in (
            (101, "Alice", "alice_tv", "777"),
            (102, "Bob", "bob_tv", "778"),
        ):
            await self.register(
                giveaway,
                telegram_id,
                name,
                f"{name.lower()}_tg",
                login,
                twitch_id,
            )
            await self.set_activity(giveaway.id, login, seconds=60, messages=1)

        await self.confirm_finish(owner)
        await self.giveaway(owner, "reroll")

        finished = await self.storage.latest_finished_giveaway()
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual(len(winners), 2)
        self.assertEqual(len({winner.telegram_user_id for winner in winners}), 2)
        notified_ids = [sent[0] for sent in self.bot.sent if sent[0] in {101, 102}]
        self.assertEqual(len(notified_ids), 2)
        self.assertEqual(set(notified_ids), {101, 102})

    async def test_notification_failure_does_not_cancel_finished_giveaway(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Недоступные личные сообщения")
        giveaway = await self.storage.active_giveaway()
        await self.register(giveaway, 101, "Alice", "alice_tg", "alice_tv", "777")
        await self.set_activity(giveaway.id, "alice_tv", seconds=60, messages=1)
        self.bot.failed_send_ids.add(101)

        await self.confirm_finish(owner)

        self.assertIsNone(await self.storage.active_giveaway())
        finished = await self.storage.latest_finished_giveaway()
        winners = await self.storage.recorded_winners(finished)
        self.assertEqual([winner.telegram_user_id for winner in winners], [101])
        combined = "\n".join(answer[0] for answer in owner.answers)
        self.assertIn("Не удалось отправить личное сообщение", combined)

    async def test_all_admin_aliases_are_registered_and_execute_expected_actions(self) -> None:
        expected = {
            "start",
            "link",
            "status",
            "giveaway_create",
            "giveaway_edit",
            "giveaway_announce_start",
            "giveaway_status",
            "giveaway_participants",
            "giveaway_finish",
            "giveaway_announce_finish",
            "giveaway_report",
            "giveaway_reroll",
            "giveaway",
            "viewers",
            "twitch_announce",
        }
        self.assertEqual({command.command for command in owner_bot_commands()}, expected)

        owner = self.message(1)
        await self.shortcut(owner, "giveaway_create", "1 1 1 0 1 Алиасы")
        self.assertIsNotNone(await self.storage.active_giveaway())
        await self.shortcut(owner, "giveaway_edit", "minutes 2")
        self.assertEqual((await self.storage.active_giveaway()).min_seconds, 120)
        await self.shortcut(owner, "twitch_announce", "on 15")
        self.assertIn("Twitch-анонсы включены", owner.last_text)
        await self.shortcut(owner, "giveaway_announce_start")
        self.assertIn("Розыгрыш начался", self.bot.sent[-1][1])
        await self.shortcut(owner, "giveaway_status")
        self.assertIn("Текущий статус", owner.last_text)
        await self.shortcut(owner, "giveaway_participants")
        self.assertIn("пока нет участников", owner.last_text)
        await self.shortcut(owner, "giveaway_finish")
        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("Подтвердите завершение", owner.last_text)
        await self.shortcut(owner, "giveaway_finish", "confirm")
        self.assertIsNone(await self.storage.active_giveaway())
        await self.shortcut(owner, "giveaway_announce_finish")
        self.assertIn("Розыгрыш завершён", self.bot.sent[-1][1])
        await self.shortcut(owner, "giveaway_report")
        self.assertEqual(self.bot.documents[-1][0], self.settings.owner_telegram_id)
        self.assertIn("XLSX-отчёт отправлен", owner.last_text)
        await self.shortcut(owner, "giveaway_reroll")
        self.assertIn("Подходящих участников не осталось", owner.last_text)

    async def test_non_owner_cannot_run_giveaway_commands_or_aliases(self) -> None:
        user = self.message(101)

        await self.giveaway(user, "start 1 1")
        await self.shortcut(user, "giveaway_create", "1 1")
        await self.shortcut(user, "giveaway_edit", "minutes 999")
        await self.shortcut(user, "giveaway_report")
        await self.shortcut(user, "twitch_announce", "on 15")

        self.assertEqual(user.answers, [])
        self.assertIsNone(await self.storage.active_giveaway())

    async def test_non_owner_cannot_edit_an_active_giveaway(self) -> None:
        giveaway = await self.storage.start_giveaway(10, 2, title="Только владелец")
        user = self.message(101)

        await self.giveaway(user, "edit minutes 999")
        await self.shortcut(user, "giveaway_edit", "messages 999")

        self.assertEqual(user.answers, [])
        unchanged = await self.storage.active_giveaway()
        self.assertEqual(unchanged.id, giveaway.id)
        self.assertEqual((unchanged.min_seconds, unchanged.min_messages), (600, 2))

    async def test_finish_confirmation_can_be_cancelled_and_cannot_be_skipped(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Подтверждение")

        await self.giveaway(owner, "finish confirm")
        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("Нет действующего подтверждения", owner.last_text)

        await self.giveaway(owner, "finish force confirm")
        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("Нет действующего подтверждения", owner.last_text)

        await self.giveaway(owner, "finish")
        await self.giveaway(owner, "finish cancel")
        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("отменено", owner.last_text)

        await self.giveaway(owner, "finish confirm")
        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("Нет действующего подтверждения", owner.last_text)

    async def test_finish_confirmation_expires_after_two_minutes(self) -> None:
        owner = self.message(1)
        await self.giveaway(owner, "start 1 1 1 0 1 Истечение подтверждения")

        with patch("app.main.time.time", side_effect=[1000.0, 1121.0]):
            await self.giveaway(owner, "finish")
            await self.giveaway(owner, "finish confirm")

        self.assertIsNotNone(await self.storage.active_giveaway())
        self.assertIn("истекли две минуты", owner.last_text)

    async def test_admin_commands_report_when_required_giveaway_is_missing(self) -> None:
        cases = (
            ("edit minutes 5", "нет активного розыгрыша для изменения"),
            ("announce_start", "Нет активного розыгрыша для анонса"),
            ("status", "Сейчас нет активного розыгрыша"),
            ("participants", "Пока нет розыгрыша"),
            ("finish", "Нет активного розыгрыша для завершения"),
            ("announce_finish", "Нет завершённого розыгрыша для анонса"),
            ("report", "Нет завершённого розыгрыша для XLSX-отчёта"),
            ("reroll", "Нет завершённого розыгрыша для перевыбора"),
        )
        for args, expected in cases:
            with self.subTest(args=args):
                owner = self.message(1)
                await self.giveaway(owner, args)
                self.assertIn(expected, owner.last_text)

    async def test_unknown_giveaway_action_returns_complete_help(self) -> None:
        owner = self.message(1)

        await self.giveaway(owner, "unknown")

        for action in (
            "start",
            "edit",
            "announce_start",
            "status",
            "participants",
            "finish",
            "announce_finish",
            "report",
            "reroll",
        ):
            self.assertIn(action, owner.last_text)


if __name__ == "__main__":
    unittest.main()
