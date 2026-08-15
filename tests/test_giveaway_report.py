from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace
import unittest

from openpyxl import load_workbook

from app.giveaway_report import (
    GiveawayReportMetadata,
    GiveawayReportParticipant,
    GiveawayReportRound,
    RESULT_NOT_ELIGIBLE,
    RESULT_NOT_SELECTED,
    RESULT_WINNER,
    build_giveaway_report,
)


class GiveawayReportTests(unittest.TestCase):
    def metadata(self, **overrides: object) -> GiveawayReportMetadata:
        values: dict[str, object] = {
            "giveaway_id": 17,
            "title": "Розыгрыш ключей 🎁",
            "prize": "Steam key",
            "min_minutes": 60,
            "min_messages": 5,
            "winner_count": 1,
            "min_participants": 2,
            "message_interval_seconds": 30,
            "started_at": 1_787_000_000,
            "end_at": 1_787_086_400,
            "finished_at": 1_787_090_000,
            "eligible_count": 2,
        }
        values.update(overrides)
        return GiveawayReportMetadata(**values)  # type: ignore[arg-type]

    @staticmethod
    def participant(
        index: int,
        *,
        admitted: bool = True,
        result: str = RESULT_NOT_SELECTED,
    ) -> GiveawayReportParticipant:
        return GiveawayReportParticipant(
            twitch_login=f"viewer_{index}",
            minutes=60 + index / 10,
            messages=5 + index,
            admitted=admitted,
            random_number=str(900_000_000_000 - index),
            eligible_place=index if admitted else None,
            result=result,
        )

    def workbook(self, data: bytes):
        self.assertTrue(data.startswith(b"PK\x03\x04"))
        return load_workbook(BytesIO(data), data_only=False)

    def test_round_trip_has_two_public_sheets_and_typed_values(self) -> None:
        participants = (
            self.participant(1, result=RESULT_WINNER),
            self.participant(2),
            self.participant(3, admitted=False, result=RESULT_NOT_ELIGIBLE),
        )
        data = build_giveaway_report(
            self.metadata(),
            (GiveawayReportRound(1, False, participants),),
        )

        workbook = self.workbook(data)
        self.assertEqual(workbook.sheetnames, ["Итоги", "Участники"])
        sheet = workbook["Участники"]
        self.assertEqual(
            [sheet.cell(row=4, column=column).value for column in range(1, 9)],
            [
                "№",
                "Twitch",
                "Минуты",
                "Сообщения",
                "Допущен",
                "Случайное число",
                "Место среди допущенных",
                "Результат",
            ],
        )
        self.assertIsInstance(sheet["C5"].value, float)
        self.assertIsInstance(sheet["D5"].value, int)
        self.assertEqual(sheet["F5"].value, 899999999999)
        self.assertEqual(sheet["F5"].data_type, "n")
        self.assertEqual(sheet["F5"].number_format, "000000000000")
        self.assertEqual(sheet["G5"].value, 1)
        self.assertEqual(sheet["H5"].value, RESULT_WINNER)
        self.assertEqual(sheet.freeze_panes, "A5")
        self.assertTrue(sheet.tables)
        self.assertRegex(workbook["Итоги"]["B8"].value, r"^\d{2}\.\d{2}\.\d{4} \d{2}:\d{2} \(МСК\)$")
        workbook.close()

    def test_zero_participants_is_a_valid_workbook(self) -> None:
        data = build_giveaway_report(
            self.metadata(eligible_count=0),
            (GiveawayReportRound(1, False, ()),),
        )

        workbook = self.workbook(data)
        sheet = workbook["Участники"]
        self.assertEqual(sheet.max_row, 4)
        self.assertTrue(sheet.tables)
        self.assertEqual(workbook["Итоги"]["A1"].value, "Публичный отчёт о розыгрыше")
        workbook.close()

    def test_one_hundred_rows_and_multiple_rounds_are_not_truncated(self) -> None:
        first_round = tuple(
            self.participant(
                index,
                result=RESULT_WINNER if index == 1 else RESULT_NOT_SELECTED,
            )
            for index in range(1, 101)
        )
        second_round = tuple(
            self.participant(
                index,
                result=RESULT_WINNER if index == 2 else RESULT_NOT_SELECTED,
            )
            for index in range(1, 101)
        )
        data = build_giveaway_report(
            self.metadata(),
            (
                GiveawayReportRound(1, False, first_round),
                GiveawayReportRound(2, True, second_round),
            ),
        )

        workbook = self.workbook(data)
        self.assertEqual(workbook.sheetnames, ["Итоги", "Раунд 1", "Раунд 2"])
        self.assertEqual(workbook["Раунд 1"].max_row, 104)
        self.assertEqual(workbook["Раунд 2"]["B104"].value, "viewer_100")
        self.assertIn("Принудительный режим: Да", workbook["Раунд 2"]["A2"].value)
        workbook.close()

    def test_unicode_is_preserved_and_formula_xml_injection_is_neutralized(self) -> None:
        metadata = SimpleNamespace(
            giveaway_id=8,
            title="=HYPERLINK(\"https://bad\")\x00 🎮",
            prize="+cmd|'/C calc'!A0\x07",
            min_minutes=1,
            min_messages=1,
            winner_count=1,
            min_participants=1,
            message_interval_seconds=0,
            started_at=None,
            end_at=None,
            finished_at=None,
            eligible_count=1,
            telegram_user_id=999999,
            telegram_username="private_tg",
        )
        participant = {
            "twitch_login": "=опасный_логин\x01",
            "minutes": 1.5,
            "messages": 2,
            "admitted": True,
            "random_number": "9223372036854775807",
            "eligible_place": 1,
            "result": RESULT_WINNER,
            "telegram_user_id": 123456,
            "telegram_name": "Совершенно секретно",
            "telegram_username": "hidden_tg",
        }
        data = build_giveaway_report(
            metadata,
            ({"draw_round": 1, "forced": False, "participants": [participant]},),
        )

        workbook = self.workbook(data)
        all_cells = [
            cell
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        all_text = "\n".join(str(cell.value) for cell in all_cells)
        self.assertIn("🎮", all_text)
        self.assertNotIn("\x00", all_text)
        self.assertNotIn("\x01", all_text)
        self.assertNotIn("\x07", all_text)
        self.assertNotIn("999999", all_text)
        self.assertNotIn("123456", all_text)
        self.assertNotIn("private_tg", all_text)
        self.assertNotIn("hidden_tg", all_text)
        self.assertNotIn("Совершенно секретно", all_text)
        self.assertTrue(all(cell.data_type != "f" for cell in all_cells))
        self.assertEqual(workbook["Участники"]["B5"].value, "=опасный_логин")
        self.assertEqual(workbook["Участники"]["B5"].data_type, "s")
        self.assertEqual(workbook["Участники"]["F5"].value, "9223372036854775807")
        self.assertNotIn("telegram", workbook.properties.description.casefold())
        workbook.close()

    def test_rejects_inconsistent_result_and_non_decimal_score(self) -> None:
        inconsistent = self.participant(
            1, admitted=False, result=RESULT_NOT_SELECTED
        )
        with self.assertRaisesRegex(ValueError, "не соответствует"):
            build_giveaway_report(
                self.metadata(),
                (GiveawayReportRound(1, False, (inconsistent,)),),
            )

        bad_score = SimpleNamespace(
            twitch_login="viewer",
            minutes=1,
            messages=1,
            admitted=True,
            random_number="9.22e18",
            eligible_place=1,
            result=RESULT_WINNER,
        )
        with self.assertRaisesRegex(ValueError, "десятичным идентификатором"):
            build_giveaway_report(
                self.metadata(),
                ({"draw_round": 1, "forced": False, "participants": [bad_score]},),
            )


if __name__ == "__main__":
    unittest.main()
